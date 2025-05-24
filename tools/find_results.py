import os
import json
from json import JSONDecoder, JSONDecodeError
import re
import argparse

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook


def dict_to_figure(num_sampling, make_dict, plt_title, save_path):
    for name, param in make_dict.copy().items():
        make_dict[name] = [k for j, k in enumerate(param) if j%num_sampling == 0]
    df = pd.DataFrame(make_dict)
    df = df.transpose()

    # fig = plt.figure()
    plt.rcParams['figure.figsize'] = [30, 16]
    # sns.heatmap(df, annot=True, fmt='d')
    # sns.heatmap(df, cmap='RdYlGn_r')
    # sns.heatmap(df, cmap='YlGnBu')  #
    plt.pcolor(df, cmap='plasma') # https://matplotlib.org/tutorials/colors/colormaps.html
    plt.xticks(np.arange(0.5, len(df.columns), 1), df.columns)
    plt.yticks(np.arange(0.5, len(df.index), 1), df.index)
    plt.title(plt_title, fontsize=20)
    plt.xlabel('BIN_params', fontsize=2)
    plt.ylabel('iter', fontsize=2)
    plt.colorbar()
    plt.savefig(save_path, dpi = dpi)
    plt.cla()
    plt.close()

def decode_stacked(document, pos=0, decoder=JSONDecoder()):
    while True:
        NOT_WHITESPACE = re.compile(r'[^\s]')
        match = NOT_WHITESPACE.search(document, pos)
        if not match:
            return
        pos = match.start()

        try:
            obj, pos = decoder.raw_decode(document, pos)
        except JSONDecodeError:
            # do something sensible if there's some error
            raise
        yield obj


if __name__ == "__main__":

    parser = argparse.ArgumentParser()
    parser.add_argument('--folder_dir', default='../logs')
    parser.add_argument('--folder_name', default='Visualize')
    parser.add_argument('--start_name', default='j02')
    parser.add_argument('--end_name', default='j02')
    parser.add_argument('--except_number', default='99')
    parser.add_argument('--max_only', default='0')
    parser.add_argument('--name_length', default='1')
    parser.add_argument('--print_none', default='0')
    parser.add_argument('--zfill', default='2')

    parser.add_argument('--market', default='0')
    parser.add_argument('--duke', default='0')
    parser.add_argument('--bin_all', default='0') # [x epoch, y layer] BIN_mean, BIN_std
    parser.add_argument('--bin_hist_img', default='0') # for layer [x epoch, y hist]
    parser.add_argument('--bin_hist_xlsx', default='0') # [new] for layer [x epoch, y hist]
    parser.add_argument('--bin_last_hist_img', default='0') # plot final histogram
    parser.add_argument('--bin_last_xlsx', default='0') # compute final histogram -> excel
    parser.add_argument('--bin_last_img', default='0') # [new] compute final histogram -> image
    parser.add_argument('--bin_stat_img', default='0') # [x epoch, y layer] BIN_mean, BIN_std
    parser.add_argument('--bin_stat_xlsx', default='0') # [new] [x epoch, y layer] BIN_mean, BIN_std
    parser.add_argument('--bin_grad_xlsx', default='0')
    parser.add_argument('--bin_loss_xlsx', default='0')
    parser.add_argument('--bin_cls_xlsx', default='0')
    parser.add_argument('--num_sampling', default='1')
    parser.add_argument('--dpi', default='300')
    parser.add_argument('--print_row', default='0')
    parser.add_argument('--map', default='0')

    args = parser.parse_args()
    print_row = int(args.print_row)
    map = int(args.map)
    folder_dir = args.folder_dir
    folder_name = args.folder_name
    start_name = args.start_name
    end_name = args.end_name
    max_only = args.max_only
    bin_stat_img = int(args.bin_stat_img)
    bin_stat_xlsx = int(args.bin_stat_xlsx)
    bin_hist_img = int(args.bin_hist_img)
    bin_hist_xlsx = int(args.bin_hist_xlsx)
    bin_last_hist_img = int(args.bin_last_hist_img)
    bin_last_xlsx = int(args.bin_last_xlsx)
    bin_last_img = int(args.bin_last_img)
    bin_grad_xlsx = int(args.bin_grad_xlsx)
    bin_loss_xlsx = int(args.bin_loss_xlsx)
    bin_cls_xlsx = int(args.bin_cls_xlsx)
    dpi = int(args.dpi)
    num_sampling = int(args.num_sampling)
    except_number = args.except_number
    except_number = [int(x) for x in except_number.split(',')]
    name_length = int(args.name_length)
    print_none = int(args.print_none)
    market = int(args.market)
    duke = int(args.duke)

    bin_all = int(args.bin_all)
    if bin_all == 1:
        bin_stat_img = 1
        bin_stat_xlsx = 1
        bin_hist_xlsx = 1
        bin_last_xlsx = 1
        bin_last_img = 1
        bin_grad_xlsx = 1
        bin_loss_xlsx = 1
        bin_cls_xlsx = 1
        bin_hist_img = 0 # slow
        bin_last_hist_img = 0 # slow
    elif bin_all == 2:
        bin_stat_img = 1
        bin_stat_xlsx = 1
        bin_hist_xlsx = 1
        bin_last_xlsx = 1
        bin_last_img = 1
        bin_grad_xlsx = 1
        bin_loss_xlsx = 1
        bin_cls_xlsx = 1
        bin_hist_img = 1 # slow
        bin_last_hist_img = 1 # slow


    txt_name = start_name[0:name_length]
    try:
        start_num = int(start_name[name_length:])
        end_num = int(end_name[name_length:])
        err_num = False
    except:
        err_num = True
        start_num = 0
        end_num = 0

    if market == 1 and duke == 0:
        acc_main_txt = "Market1501/Rank-1"
        acc_txt = []
        acc_txt.append("Market1501/Rank-1")
        acc_txt.append("DukeMTMC/Rank-1")
        acc_txt.append("iteration")
    elif market == 0 and duke == 1:
        acc_main_txt = "DukeMTMC/Rank-1"
        acc_txt = []
        acc_txt.append("Market1501/Rank-1")
        acc_txt.append("DukeMTMC/Rank-1")
        acc_txt.append("iteration")
    else:
        if map:
            acc_main_txt = "** all_average **/mAP"
            # acc_main_txt = "** all_average **/Rank-1"
        else:
            acc_main_txt = "** all_average **/Rank-1"
        acc_txt = []
        acc_txt.append("** all_average **/Rank-1")
        acc_txt.append("ALL_GRID_average/Rank-1")
        acc_txt.append("ALL_PRID_average/Rank-1")
        acc_txt.append("ALL_VIPER_only_10_average/Rank-1")
        acc_txt.append("ALL_iLIDS_average/Rank-1")
        acc_txt.append("** all_average **/mAP")
        acc_txt.append("ALL_GRID_average/mAP")
        acc_txt.append("ALL_PRID_average/mAP")
        acc_txt.append("ALL_VIPER_only_10_average/mAP")
        acc_txt.append("ALL_iLIDS_average/mAP")
        acc_txt.append("iteration")

    bin_main_txt = "_g_mean"
    bin_mean = dict()
    bin_std = dict()
    bin_iter = []
    bin_folder_name = "BIN_params"
    bin_folder_path = os.path.join(folder_dir, folder_name, bin_folder_name)
    if not os.path.isdir(bin_folder_path):
        os.mkdir(bin_folder_path)

    for i in range(start_num, end_num + 1):
        if err_num:
            case = txt_name
        else:
            case = txt_name + str(i).zfill(int(args.zfill))
        file_name = os.path.join(folder_dir, folder_name, case, 'metrics.json')
        # print(case)
        if os.path.isfile(file_name) and i not in except_number:

            file = open(file_name, "r", encoding='utf-8')
            ##########################################################################################
            # BIN histogram
            ##########################################################################################
            if bin_hist_img == 1 or bin_hist_xlsx == 1 or bin_last_hist_img == 1 or bin_last_xlsx == 1 or bin_last_img == 1:

                # for bin_gate parameters
                cnt = 0
                for obj in decode_stacked(file.read()):

                    if cnt == 0:
                        all_name = []
                        for name, param in obj.items():
                            if 'b4' in name:
                                name = name.replace('b4', 'b')
                            if '_g_mean' in name:
                                all_name.append(name.replace('_g_mean', ''))
                        all_hist = dict()
                        for j in range(len(all_name)):
                            all_hist[all_name[j]] = dict()
                            for k in range(20):
                                all_hist[all_name[j]][str(k)] = []

                    local_hist = dict()
                    for j in range(len(all_name)):
                        local_hist[all_name[j]] = np.zeros(20)

                    # local_hist = np.zeros(20)
                    for name, val in obj.items():
                        if 'b4' in name:
                            name = name.replace('b4', 'b')
                        if '_g_hist' in name:
                            hist_target = name[:name.find('_g_hist')]
                            hist_number = int(name[name.find('_g_hist')+7:])
                            local_hist[hist_target][hist_number] = val

                    for name, val in local_hist.items():
                        local_hist[name] /= np.sum(local_hist[name])
                        for k in range(20):
                            all_hist[name][str(k)].append(local_hist[name][k])

                    cnt += 1

                if bin_last_hist_img == 1 or bin_last_xlsx == 1 or bin_last_img == 1:
                    final_hist = dict()
                    for j in range(len(all_name)):
                        final_hist[all_name[j]] = np.zeros(20)
                    for name, val in all_hist.items():
                        for k in range(20):
                            final_hist[name][k] = val[str(k)][-2]

                    if bin_last_hist_img == 1:
                        for name, val in all_hist.items():
                            if not os.path.isdir(os.path.join(bin_folder_path, case + '_hist_final')):
                                os.mkdir(os.path.join(bin_folder_path, case + '_hist_final'))
                            save_path = os.path.join(bin_folder_path, case + '_hist_final', name + '.png')
                            plt.rcParams['figure.figsize'] = [15, 15]
                            plt.bar(np.arange(20), final_hist[name], color='green', alpha=0.5)
                            plt.grid()
                            plt.xlabel('Weight', fontsize=14)
                            plt.title(name, fontsize=14)
                            plt.xticks(np.arange(20), [x/20 for x in range(20)], fontsize=14)
                            plt.yticks(fontsize=14)

                            plt.savefig(save_path, dpi=dpi)
                            plt.cla()
                            plt.close()
                    if bin_last_xlsx == 1:
                        wb = Workbook()
                        sheet1 = wb.active
                        sheet1.title = 'bin_parameter'
                        row_index = [x+1 for x in range(len(final_hist))]
                        cnt = 0
                        for name, val in final_hist.items():
                            sheet1.cell(row=int(row_index[cnt]), column=1).value = name
                            for k in range(len(val)):
                                sheet1.cell(row=int(row_index[cnt]), column=k+2).value = val[k]
                            cnt += 1
                        wb.save(filename=os.path.join(bin_folder_path, case + '_last_bin.xlsx'))
                    if bin_last_img == 1:
                        save_path = os.path.join(bin_folder_path, case + '_last_bin.png')
                        dict_to_figure(1, final_hist, 'BIN histogram (last epoch)', save_path)

                if bin_hist_img == 1:
                    for name, val in all_hist.items():
                        if not os.path.isdir(os.path.join(bin_folder_path, case + '_hist')):
                            os.mkdir(os.path.join(bin_folder_path, case + '_hist'))
                        save_path = os.path.join(bin_folder_path, case + '_hist', name + '.png')
                        dict_to_figure(num_sampling, val, 'histogram', save_path)
                if bin_hist_xlsx == 1:
                    wb = Workbook()
                    cnt_case = 0
                    sheet1 = wb.active
                    sheet1.title = 'all_hist'
                    for name_all, val_all in all_hist.items():
                        row_index = [x+1+cnt_case for x in range(len(val_all))]
                        cnt = 0
                        for name, val in val_all.items():
                            sheet1.cell(row=int(row_index[cnt]), column=1).value = name_all + '_' + name
                            for k in range(len(val)):
                                sheet1.cell(row=int(row_index[cnt]), column=k+2).value = val[k]
                            cnt += 1
                        cnt_case = max(row_index) + 1
                    wb.save(filename=os.path.join(bin_folder_path, case + '_all_hist.xlsx'))

            ##########################################################################################
            # BIN mean, std
            ##########################################################################################
            file = open(file_name, "r", encoding='utf-8')
            if bin_stat_img == 1 or bin_stat_xlsx == 1:
                # for bin_gate parameters
                bin_iter = []
                bin_dict = dict()
                cnt = 0
                for obj in decode_stacked(file.read()):
                    for name, param in obj.items():
                        if 'b4' in name:
                            name = name.replace('b4', 'b')
                        if '_g_mean' in name:
                            if cnt == 0:
                                bin_mean[name] = []
                                bin_mean[name].append(param)
                            else:
                                bin_mean[name].append(param)
                        if '_g_std' in name:
                            if cnt == 0:
                                bin_std[name] = []
                                bin_std[name].append(param)
                            else:
                                bin_std[name].append(param)
                    cnt += 1
                    bin_iter.append(obj['iteration'])

                if bin_stat_img == 1:
                    save_path = os.path.join(bin_folder_path, case + '_BIN_mean.png')
                    dict_to_figure(num_sampling, bin_mean, 'BIN_parameters (mean)', save_path)
                    save_path = os.path.join(bin_folder_path, case + '_BIN_std.png')
                    dict_to_figure(num_sampling, bin_std, 'BIN_parameters (std)', save_path)
                if bin_stat_xlsx == 1:
                    wb = Workbook()
                    sheet1 = wb.active
                    sheet1.title = 'BIN_mean'
                    row_index = [x+1 for x in range(len(bin_mean))]
                    cnt = 0
                    for name, val in bin_mean.items():
                        sheet1.cell(row=int(row_index[cnt]), column=1).value = name
                        for k in range(len(val)):
                            sheet1.cell(row=int(row_index[cnt]), column=k+2).value = val[k]
                        cnt += 1
                    wb.save(filename=os.path.join(bin_folder_path, case + '_BIN_mean.xlsx'))

                    wb = Workbook()
                    sheet1 = wb.active
                    sheet1.title = 'BIN_std'
                    row_index = [x + 1 for x in range(len(bin_std))]
                    cnt = 0
                    for name, val in bin_std.items():
                        sheet1.cell(row=int(row_index[cnt]), column=1).value = name
                        for k in range(len(val)):
                            sheet1.cell(row=int(row_index[cnt]), column=k + 2).value = val[k]
                        cnt += 1
                    wb.save(filename=os.path.join(bin_folder_path, case + '_BIN_std.xlsx'))

            ##########################################################################################
            # grad & loss & cls_accuracy
            ##########################################################################################
            file = open(file_name, "r", encoding='utf-8')
            if bin_grad_xlsx == 1 or bin_cls_xlsx == 1 or bin_loss_xlsx == 1:
                # for bin_gate parameters
                local_loss = dict()
                local_grad_average = dict()
                local_grad_prob = dict()

                bin_iter = []
                bin_cls = []

                cnt = 0
                for obj in decode_stacked(file.read()):
                    # for name, param in obj.items():

                    bin_cls.append(obj['cls_accuracy'])
                    bin_iter.append(obj['iteration'])

                    if cnt == 0:
                        all_name_loss = []
                        for name, param in obj.items():
                            if 'loss' in name:
                                all_name_loss.append(name)

                        all_name_layer = []
                        for name, param in obj.items():
                            if 'grad_average' in name:
                                all_name_layer.append(name.replace('_grad_average', ''))

                        for kk in all_name_loss:
                            local_loss[kk] = []

                        for kk in all_name_layer:
                            local_grad_average[kk] = []
                            local_grad_prob[kk] = []


                    for name, val in obj.items():
                        if 'grad_average' in name:
                            for kk in all_name_layer:
                                if kk in name:
                                    local_grad_average[kk].append(val)
                        if 'grad_prob' in name:
                            for kk in all_name_layer:
                                if kk in name:
                                    local_grad_prob[kk].append(val)
                        if 'loss' in name:
                            for kk in all_name_loss:
                                if kk in name:
                                    local_loss[kk].append(val)
                    cnt += 1
                if bin_grad_xlsx == 1:
                    wb = Workbook()
                    sheet1 = wb.active
                    sheet1.title = 'grad_average'
                    row_index = [x+1 for x in range(len(local_grad_average))]
                    cnt = 0
                    for name, val in local_grad_average.items():
                        sheet1.cell(row=int(row_index[cnt]), column=1).value = name
                        for k in range(len(val)):
                            sheet1.cell(row=int(row_index[cnt]), column=k+2).value = val[k]
                        cnt += 1
                    wb.save(filename=os.path.join(bin_folder_path, case + '_grad_average.xlsx'))

                    wb = Workbook()
                    sheet1 = wb.active
                    sheet1.title = 'grad_prob'
                    row_index = [x+1 for x in range(len(local_grad_prob))]
                    cnt = 0
                    for name, val in local_grad_prob.items():
                        sheet1.cell(row=int(row_index[cnt]), column=1).value = name
                        for k in range(len(val)):
                            sheet1.cell(row=int(row_index[cnt]), column=k+2).value = val[k]
                        cnt += 1
                    wb.save(filename=os.path.join(bin_folder_path, case + '_grad_prob.xlsx'))

                if bin_cls_xlsx == 1:
                    wb = Workbook()
                    sheet1 = wb.active
                    sheet1.title = 'cls'
                    sheet1.cell(row=1, column=1).value = 'iter'
                    for k in range(len(bin_iter)):
                        sheet1.cell(row=1, column=k+2).value = bin_iter[k]
                    sheet1.cell(row=2, column=1).value = 'cls'
                    for k in range(len(bin_cls)):
                        sheet1.cell(row=2, column=k+2).value = bin_cls[k]
                    wb.save(filename=os.path.join(bin_folder_path, case + '_cls.xlsx'))
                if bin_loss_xlsx == 1:
                    wb = Workbook()
                    sheet1 = wb.active
                    sheet1.title = 'loss'
                    row_index = [x+1 for x in range(len(local_loss))]
                    cnt = 0
                    for name, val in local_loss.items():
                        sheet1.cell(row=int(row_index[cnt]), column=1).value = name
                        for k in range(len(val)):
                            sheet1.cell(row=int(row_index[cnt]), column=k+2).value = val[k]
                        cnt += 1
                    wb.save(filename=os.path.join(bin_folder_path, case + '_loss.xlsx'))



            ##########################################################################################
            # Accuracy
            ##########################################################################################
            file = open(file_name, "r", encoding='utf-8')
            # for accuracy
            acc_dict = dict()
            for obj in decode_stacked(file.read()):
                if acc_main_txt in obj.keys():
                    for find_name in acc_txt:
                        if find_name in obj.keys():
                            if find_name in acc_dict.keys():
                                acc_dict[find_name].append(obj[find_name])
                            else:
                                acc_dict[find_name] = []
                                acc_dict[find_name].append(obj[find_name])
            if len(acc_dict) > 0:
                if max_only == '1':
                    max_val = max(acc_dict[acc_main_txt])
                    max_idx = [i for i, x in enumerate(acc_dict[acc_main_txt]) if x == max_val][0]
                    try: acc_G = acc_dict["ALL_GRID_average/Rank-1"][max_idx]
                    except: acc_G = 0
                    try: acc_P = acc_dict["ALL_PRID_average/Rank-1"][max_idx]
                    except: acc_P = 0
                    try: acc_V = acc_dict["ALL_VIPER_only_10_average/Rank-1"][max_idx]
                    except: acc_V = 0
                    try: acc_I = acc_dict["ALL_iLIDS_average/Rank-1"][max_idx]
                    except: acc_I = 0
                    try: acc_M = acc_dict["Market1501/Rank-1"][max_idx]
                    except: acc_M = 0
                    try: acc_D = acc_dict["DukeMTMC/Rank-1"][max_idx]
                    except: acc_D = 0
                    try: map_all = acc_dict["** all_average **/mAP"][max_idx]
                    except: map_all = 0
                    try: map_G = acc_dict["ALL_GRID_average/mAP"][max_idx]
                    except: map_G = 0
                    try: map_P = acc_dict["ALL_PRID_average/mAP"][max_idx]
                    except: map_P = 0
                    try: map_V = acc_dict["ALL_VIPER_only_10_average/mAP"][max_idx]
                    except: map_V = 0
                    try: map_I = acc_dict["ALL_iLIDS_average/mAP"][max_idx]
                    except: map_I = 0
                    try: map_M = acc_dict["Market1501/mAP"][max_idx]
                    except: map_M = 0
                    try: map_D = acc_dict["DukeMTMC/mAP"][max_idx]
                    except: map_D = 0
                    print('[{}], final_iter:{}, eta:{}hours (iter:{}, avg:{}/{}, V:{}/{}, P:{}/{}, G:{}/{}, I:{}/{}, M:{}/{}, D:{}/{})'.
                          format(case,
                                 round(obj['iteration'] / min(acc_dict['iteration']), 2),
                                 round(obj['eta_seconds'] / 3600, 1),
                                 round(acc_dict["iteration"][max_idx] / min(acc_dict['iteration']), 0),
                                 round(max_val * 100, 2),
                                 round(map_all * 100, 2),
                                 round(acc_V * 100, 2),
                                 round(map_V * 100, 2),
                                 round(acc_P * 100, 2),
                                 round(map_P * 100, 2),
                                 round(acc_G * 100, 2),
                                 round(map_G * 100, 2),
                                 round(acc_I * 100, 2),
                                 round(map_I * 100, 2),
                                 round(acc_M * 100, 2),
                                 round(map_M * 100, 2),
                                 round(acc_D * 100, 2),
                                 round(map_D * 100, 2),))
                else:
                    all_txt = '[{}]'.format(case)
                    unique_iteration = sorted(list(set(acc_dict['iteration'])))
                    if print_row:
                        print(all_txt)
                    for i, x in enumerate(unique_iteration):
                        idx = [j for j in range(len(acc_dict["iteration"])) if x == acc_dict["iteration"][j]]
                        main_acc = [acc_dict[acc_main_txt][m] for m in range(len(acc_dict[acc_main_txt])) if m in idx]
                        max_idx = [n for n in range(len(main_acc)) if max(main_acc) == main_acc[n]]
                        final_idx = idx[max_idx[0]]

                        if print_row:
                            print('{}'.format("%2.2f"%round(acc_dict[acc_main_txt][final_idx] * 100, 2)))
                        else:
                            all_txt += ' ({}){}'.format(int(round(acc_dict["iteration"][final_idx] / min(acc_dict['iteration']), 0)),
                                                "%2.2f"%round(acc_dict[acc_main_txt][final_idx] * 100, 2))

                    if not print_row:
                        print(all_txt)
            else:
                if print_none == 1:
                    print('[{}] not exist'.format(case))
        else:
            if i not in except_number:
                if print_none == 1:
                    print('[{}] not exist'.format(case))

